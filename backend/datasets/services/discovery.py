from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from datasets.models import DatasetSheet, FileType


def discover_and_create_sheets(dataset: Dataset) -> list[DatasetSheet]:
    """Create DatasetSheet rows from uploaded file. Returns created sheets."""
    path = Path(dataset.file.path)
    if dataset.file_type == FileType.CSV:
        sheet = DatasetSheet.objects.create(
            dataset=dataset,
            name="data",
            order=0,
            selected=True,
        )
        return [sheet]

    if dataset.file_type == FileType.XLSX:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            names = list(wb.sheetnames)
        finally:
            wb.close()
        sheets: list[DatasetSheet] = []
        for order, name in enumerate(names):
            sheets.append(
                DatasetSheet.objects.create(
                    dataset=dataset,
                    name=name,
                    order=order,
                    selected=(order == 0),
                )
            )
        return sheets

    return []
