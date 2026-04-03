from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from datasets.models import FileType, InferredType


@dataclass(frozen=True)
class ColumnProfileRow:
    column_name: str
    inferred_type: str
    null_ratio: float
    sample_values: list[Any]


@dataclass(frozen=True)
class SheetProfileResult:
    row_count: int
    column_count: int
    columns: list[ColumnProfileRow]


def _try_float(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    try:
        float(s.replace(",", ""))
        return True
    except ValueError:
        return False


_DATE_RES = [
    re.compile(r"^\d{4}-\d{2}-\d{2}$"),
    re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$"),
    re.compile(r"^\d{1,2}/\d{1,2}/\d{4}$"),
]


def _try_date_str(s: str) -> bool:
    s = s.strip()
    if not s:
        return False
    for r in _DATE_RES:
        if r.match(s):
            return True
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            datetime.strptime(s, fmt)
            return True
        except ValueError:
            continue
    return False


def _classify_cell(value: Any) -> str:
    if value is None or value == "":
        return InferredType.UNKNOWN
    if isinstance(value, (int, float, Decimal)):
        return InferredType.NUMBER
    if isinstance(value, (date, datetime)):
        return InferredType.DATE
    s = str(value).strip()
    if not s:
        return InferredType.UNKNOWN
    if _try_float(s):
        return InferredType.NUMBER
    if _try_date_str(s):
        return InferredType.DATE
    return InferredType.STRING


def _finalize_types(counts: dict[str, dict[str, int]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for col, c in counts.items():
        total = sum(c.values())
        if total == 0:
            out[col] = InferredType.UNKNOWN
            continue
        best = max(c.items(), key=lambda x: x[1])[0]
        out[col] = best
    return out


def _sample_values(non_null: list[str], limit: int = 5) -> list[Any]:
    seen: list[Any] = []
    for v in non_null:
        if v not in seen:
            seen.append(v)
        if len(seen) >= limit:
            break
    return seen


def profile_csv_full(path: Path) -> SheetProfileResult:
    """Profile CSV using a bounded sample for types; row_count is exact."""
    encodings = ("utf-8-sig", "utf-8", "cp932")
    for enc in encodings:
        try:
            with path.open(newline="", encoding=enc) as f:
                reader = csv.reader(f)
                header = next(reader)
                header = [h.strip() or f"column_{i}" for i, h in enumerate(header)]
                column_count = len(header)
                col_values: dict[str, list[str]] = {h: [] for h in header}
                col_nulls: dict[str, int] = defaultdict(int)
                row_count = 0
                sampled = 0
                sample_cap = 5000
                for row in reader:
                    row_count += 1
                    if sampled < sample_cap:
                        for i, name in enumerate(header):
                            cell = row[i] if i < len(row) else ""
                            s = str(cell).strip()
                            if s == "":
                                col_nulls[name] += 1
                            else:
                                col_values[name].append(s)
                        sampled += 1
                type_counts: dict[str, dict[str, int]] = {
                    h: defaultdict(int) for h in header
                }
                for name in header:
                    for v in col_values[name]:
                        t = _classify_cell(v)
                        type_counts[name][t] += 1
                    for _ in range(col_nulls[name]):
                        type_counts[name][InferredType.UNKNOWN] += 1
                inferred = _finalize_types(type_counts)
                columns: list[ColumnProfileRow] = []
                for name in header:
                    non_null = col_values[name]
                    null_ratio = col_nulls[name] / sampled if sampled else 1.0
                    columns.append(
                        ColumnProfileRow(
                            column_name=name,
                            inferred_type=inferred[name],
                            null_ratio=min(1.0, max(0.0, null_ratio)),
                            sample_values=_sample_values(non_null),
                        )
                    )
                return SheetProfileResult(
                    row_count=row_count,
                    column_count=column_count,
                    columns=columns,
                )
        except UnicodeDecodeError:
            continue
    raise RuntimeError("CSV の文字コードを判別できませんでした。")


def profile_xlsx(path: Path, sheet_name: str) -> SheetProfileResult:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return SheetProfileResult(row_count=0, column_count=0, columns=[])
        header = [
            (str(c).strip() if c is not None else "") or f"column_{i}"
            for i, c in enumerate(header_row)
        ]
        column_count = len(header)
        col_values: dict[str, list[str]] = {h: [] for h in header}
        col_nulls: dict[str, int] = defaultdict(int)
        row_count = 0
        sampled = 0
        sample_cap = 5000
        for row in rows_iter:
            row_count += 1
            if sampled < sample_cap:
                for i, name in enumerate(header):
                    cell = row[i] if i < len(row) else None
                    if cell is None or cell == "":
                        col_nulls[name] += 1
                    else:
                        col_values[name].append(str(cell).strip())
                sampled += 1
        type_counts: dict[str, dict[str, int]] = {h: defaultdict(int) for h in header}
        for name in header:
            for v in col_values[name]:
                t = _classify_cell(v)
                type_counts[name][t] += 1
            for _ in range(col_nulls[name]):
                type_counts[name][InferredType.UNKNOWN] += 1
        inferred = _finalize_types(type_counts)
        columns: list[ColumnProfileRow] = []
        for name in header:
            non_null = col_values[name]
            null_ratio = col_nulls[name] / sampled if sampled else 1.0
            columns.append(
                ColumnProfileRow(
                    column_name=name,
                    inferred_type=inferred[name],
                    null_ratio=min(1.0, max(0.0, null_ratio)),
                    sample_values=_sample_values(non_null),
                )
            )
        return SheetProfileResult(
            row_count=row_count,
            column_count=column_count,
            columns=columns,
        )
    finally:
        wb.close()


def profile_file(
    file_type: str,
    path: Path,
    sheet_name: str | None,
) -> SheetProfileResult:
    if file_type == FileType.CSV:
        return profile_csv_full(path)
    if file_type == FileType.XLSX and sheet_name:
        return profile_xlsx(path, sheet_name)
    raise ValueError("unsupported file type or missing sheet")
