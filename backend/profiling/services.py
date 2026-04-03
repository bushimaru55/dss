from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import pandera as pa
from openpyxl import load_workbook
from pandera.errors import SchemaErrors

from datasets.models import FileType, InferredType

MISSING_TOKENS = {"", "-", "N/A", "n/a", "null", "NULL", "なし"}


@dataclass(frozen=True)
class PreviewResult:
    columns: list[str]
    rows: list[dict[str, Any]]
    summary: dict[str, Any]


@dataclass(frozen=True)
class ColumnProfile:
    original_name: str
    normalized_name: str
    inferred_dtype: str
    null_ratio: float
    unique_ratio: float
    sample_values: list[Any]
    warnings: list[str]


@dataclass(frozen=True)
class ProfilingResult:
    rows_count: int
    columns_count: int
    detected_header_row: int
    detected_data_start_row: int
    sheet_analysis: dict[str, Any]
    columns: list[ColumnProfile]


class TabularReader:
    def read_preview(
        self,
        path: Path,
        file_type: str,
        sheet_name: str | None,
        rows: int,
        header_row_1based: int | None = None,
    ) -> PreviewResult:
        raise NotImplementedError

    def read_dataframe(
        self,
        path: Path,
        file_type: str,
        sheet_name: str | None,
        header_row_1based: int | None = None,
    ) -> pd.DataFrame:
        raise NotImplementedError


class PandasTabularReader(TabularReader):
    def read_preview(
        self,
        path: Path,
        file_type: str,
        sheet_name: str | None,
        rows: int,
        header_row_1based: int | None = None,
    ) -> PreviewResult:
        df = self.read_dataframe(path, file_type, sheet_name, header_row_1based=header_row_1based).head(rows)
        df = normalize_values(df)
        summary = {}
        if file_type == FileType.XLSX and sheet_name:
            summary = analyze_excel_sheet(path, sheet_name)
        return PreviewResult(
            columns=[str(c) for c in df.columns],
            rows=df.fillna("").to_dict(orient="records"),
            summary=summary,
        )

    def read_dataframe(
        self,
        path: Path,
        file_type: str,
        sheet_name: str | None,
        header_row_1based: int | None = None,
    ) -> pd.DataFrame:
        if file_type == FileType.CSV:
            last_error: Exception | None = None
            header_idx = 0 if header_row_1based is None else max(0, header_row_1based - 1)
            for enc in ("utf-8-sig", "utf-8", "cp932"):
                try:
                    return pd.read_csv(path, encoding=enc, dtype=str, header=header_idx)
                except UnicodeDecodeError as exc:
                    last_error = exc
            raise RuntimeError("CSV の文字コードを判別できませんでした") from last_error
        if file_type == FileType.XLSX:
            if not sheet_name:
                raise ValueError("sheet_name is required for xlsx")
            header_row = (
                int(header_row_1based)
                if header_row_1based is not None
                else detect_excel_header_row(path, sheet_name)
            )
            df = pd.read_excel(
                path,
                sheet_name=sheet_name,
                engine="openpyxl",
                dtype=str,
                header=max(0, header_row - 1),
            )
            return backfill_amount_column_from_qty_unit(df)
        raise ValueError("unsupported file type")


def normalize_column_name(name: str, seen: dict[str, int]) -> str:
    value = str(name or "").replace("\n", " ").replace("\u3000", " ").strip()
    if value.lower().startswith("unnamed:") or not value:
        value = "column"
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^0-9A-Za-z_ぁ-んァ-ン一-龥]", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        value = "column"
    base = value.lower()
    seen[base] = seen.get(base, 0) + 1
    if seen[base] > 1:
        return f"{base}_{seen[base]}"
    return base


def normalize_values(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        out[col] = out[col].map(lambda v: None if str(v).strip() in MISSING_TOKENS else v)
    return out


def _series_all_empty_or_nan(series: pd.Series) -> bool:
    if series.isna().all():
        return True
    s2 = series.astype(str).str.strip()
    return bool(s2.isin(("", "nan", "None", "NaT", "<NA>")).all())


def backfill_amount_column_from_qty_unit(df: pd.DataFrame) -> pd.DataFrame:
    """openpyxl/pandas が数式セルを NaN で読む（ブックに計算結果キャッシュがない）場合に、数量×単価で金額列を埋める。"""
    out = df.copy()
    triples = (
        ("売上金額(円)", "販売数量", "単価(円)"),
        ("売上金額", "販売数量", "単価(円)"),
        ("売上金額(円)", "数量", "単価(円)"),
    )
    for amt_col, q_col, u_col in triples:
        if amt_col not in out.columns or q_col not in out.columns or u_col not in out.columns:
            continue
        if not _series_all_empty_or_nan(out[amt_col]):
            continue
        q = pd.to_numeric(out[q_col].astype(str).str.replace(",", "", regex=False), errors="coerce")
        u = pd.to_numeric(out[u_col].astype(str).str.replace(",", "", regex=False), errors="coerce")
        prod = q * u

        def _fmt(x: Any) -> str:
            if pd.isna(x):
                return ""
            xf = float(x)
            if abs(xf - round(xf)) < 1e-9:
                return str(int(round(xf)))
            return str(round(xf, 2))

        out[amt_col] = prod.map(_fmt)
    return out


def infer_dtype(series: pd.Series) -> str:
    s = series.dropna().astype(str).head(200)
    if s.empty:
        return InferredType.UNKNOWN
    numeric_like = s.str.replace(",", "", regex=False).str.replace("¥", "", regex=False).str.replace("円", "", regex=False)
    numeric_ratio = pd.to_numeric(numeric_like, errors="coerce").notna().mean()
    if numeric_ratio >= 0.8:
        return InferredType.NUMBER
    date_ratio = pd.to_datetime(s, errors="coerce").notna().mean()
    if date_ratio >= 0.8:
        return InferredType.DATE
    return InferredType.STRING


def _is_data_like_token(value: str) -> bool:
    s = (value or "").strip()
    if not s:
        return False
    s_num = s.replace(",", "").replace("¥", "").replace("円", "")
    try:
        float(s_num)
        return True
    except ValueError:
        pass
    try:
        pd.to_datetime(s, errors="raise")
        return True
    except Exception:
        return False


def read_raw_tabular_grid(
    path: Path,
    file_type: str,
    sheet_name: str | None,
    nrows: int,
) -> tuple[list[list[str]], int]:
    """Excel/CSV を header=None で読み、画面上の並びのままのグリッドを返す。各行は文字列セルのリスト。"""
    nrows = max(1, min(int(nrows), 200))
    if file_type == FileType.CSV:
        last_error: Exception | None = None
        raw = None
        for enc in ("utf-8-sig", "utf-8", "cp932"):
            try:
                raw = pd.read_csv(path, encoding=enc, header=None, nrows=nrows, dtype=str).fillna("")
                break
            except UnicodeDecodeError as exc:
                last_error = exc
        if raw is None:
            raise RuntimeError("CSV の文字コードを判別できませんでした") from last_error
    elif file_type == FileType.XLSX:
        if not sheet_name:
            raise ValueError("sheet_name is required for xlsx")
        raw = pd.read_excel(
            path,
            sheet_name=sheet_name,
            engine="openpyxl",
            header=None,
            nrows=nrows,
            dtype=str,
        ).fillna("")
    else:
        raise ValueError("unsupported file type")
    rows_out: list[list[str]] = []
    for i in range(len(raw)):
        row = [str(x) for x in raw.iloc[i].tolist()]
        rows_out.append(row)
    max_cols = max((len(r) for r in rows_out), default=0)
    for r in rows_out:
        while len(r) < max_cols:
            r.append("")
    return rows_out, max_cols


def detect_excel_header_row(path: Path, sheet_name: str, scan_rows: int = 40) -> int:
    raw = pd.read_excel(
        path,
        sheet_name=sheet_name,
        engine="openpyxl",
        header=None,
        nrows=scan_rows,
        dtype=str,
    ).fillna("")
    if raw.empty:
        return 1

    best_row = 1
    best_score = float("-inf")
    max_cols = max(1, len(raw.columns))
    metadata_keywords = ("作成日", "対象期間", "明細", "件数")

    for idx in range(len(raw)):
        row_vals = [str(v).strip() for v in raw.iloc[idx].tolist()]
        non_empty_vals = [v for v in row_vals if v]
        non_empty = len(non_empty_vals)
        if non_empty < 2:
            continue

        text_like = sum(1 for v in non_empty_vals if not _is_data_like_token(v))
        text_ratio = text_like / max(1, non_empty)
        filled_ratio = non_empty / max_cols

        next_filled_ratio = 0.0
        if idx + 1 < len(raw):
            next_vals = [str(v).strip() for v in raw.iloc[idx + 1].tolist()]
            next_filled_ratio = sum(1 for v in next_vals if v) / max_cols

        metadata_penalty = 0.0
        row_joined = " ".join(non_empty_vals)
        if any(k in row_joined for k in metadata_keywords):
            metadata_penalty = 2.0

        score = (
            filled_ratio * 8.0
            + text_ratio * 4.0
            + min(1.0, next_filled_ratio) * 3.0
            - metadata_penalty
        )
        if score > best_score:
            best_score = score
            best_row = idx + 1  # 1-based row index

    return max(1, best_row)


def analyze_excel_sheet(path: Path, sheet_name: str) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        ws = wb[sheet_name]
        merged_cells_count = len(ws.merged_cells.ranges)
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        blank_rows = 0
        header_candidate = detect_excel_header_row(path, sheet_name)
        for r in range(1, min(max_row, 30) + 1):
            row_values = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            non_empty = sum(1 for v in row_values if v not in (None, ""))
            if non_empty == 0:
                blank_rows += 1
        return {
            "merged_cells_count": merged_cells_count,
            "blank_row_ratio": (blank_rows / max(1, min(max_row, 30))),
            "detected_header_row": header_candidate,
            "possible_multi_header": merged_cells_count > 0 and header_candidate > 1,
            "table_like": max_col >= 2 and max_row >= 2,
        }
    finally:
        wb.close()


def is_structure_ambiguous(
    sheet_analysis: dict[str, Any] | None,
    df: pd.DataFrame,
    file_type: str,
) -> tuple[bool, list[str]]:
    """自動判定だけでは列名行・表の始まりに自信が持てない場合に True。"""
    reasons: list[str] = []
    sa = sheet_analysis or {}
    if file_type == FileType.XLSX:
        if sa.get("possible_multi_header"):
            reasons.append("possible_multi_header")
        if int(sa.get("merged_cells_count") or 0) > 0:
            reasons.append("merged_cells")
        br = float(sa.get("blank_row_ratio") or 0)
        if br > 0.25:
            reasons.append("high_blank_rows")
    unnamed = sum(1 for c in df.columns if str(c).lower().startswith("unnamed:"))
    if unnamed >= 2:
        reasons.append("many_unnamed_columns")
    return (len(reasons) > 0, reasons)


def build_profile(df: pd.DataFrame, sheet_analysis: dict[str, Any] | None = None) -> ProfilingResult:
    df = normalize_values(df)
    seen: dict[str, int] = {}
    normalized_names = [normalize_column_name(str(c), seen) for c in df.columns]
    result_cols: list[ColumnProfile] = []
    for original, normalized in zip(df.columns, normalized_names, strict=False):
        series = df[original]
        null_ratio = float(series.isna().mean()) if len(series) else 1.0
        non_null = series.dropna().astype(str)
        unique_ratio = float(non_null.nunique() / max(1, len(non_null)))
        warnings: list[str] = []
        if str(original).lower().startswith("unnamed:"):
            warnings.append("unnamed_header")
        if unique_ratio == 1.0 and len(non_null) > 50:
            warnings.append("high_cardinality")
        if non_null.str.contains(r"\s*/\s*", regex=True).mean() >= 0.5 if len(non_null) else False:
            warnings.append("multi_value_cell_candidate")
        result_cols.append(
            ColumnProfile(
                original_name=str(original),
                normalized_name=normalized,
                inferred_dtype=infer_dtype(series),
                null_ratio=null_ratio,
                unique_ratio=unique_ratio,
                sample_values=non_null.head(5).tolist(),
                warnings=warnings,
            )
        )
    detected_header_row = int((sheet_analysis or {}).get("detected_header_row") or 1)
    return ProfilingResult(
        rows_count=len(df),
        columns_count=len(df.columns),
        detected_header_row=detected_header_row,
        detected_data_start_row=detected_header_row + 1,
        sheet_analysis=sheet_analysis or {},
        columns=result_cols,
    )


def validate_with_pandera(df: pd.DataFrame, schema_name: str = "generic") -> dict[str, Any]:
    schemas: dict[str, pa.DataFrameSchema] = {
        "generic": pa.DataFrameSchema({}, strict=False, coerce=False),
        "sales_basic": pa.DataFrameSchema(
            {
                "amount": pa.Column(float, nullable=True, required=False),
                "date": pa.Column(str, nullable=True, required=False),
            },
            strict=False,
            coerce=False,
        ),
        "customer_basic": pa.DataFrameSchema(
            {
                "customer": pa.Column(str, nullable=True, required=False),
                "email": pa.Column(str, nullable=True, required=False),
            },
            strict=False,
            coerce=False,
        ),
    }
    schema = schemas.get(schema_name, schemas["generic"])
    try:
        schema.validate(df, lazy=True)
        return {"ok": True, "errors": []}
    except SchemaErrors as exc:
        failures = exc.failure_cases.fillna("").to_dict(orient="records")
        return {"ok": False, "errors": failures[:100]}
