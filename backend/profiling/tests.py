from pathlib import Path

from openpyxl import Workbook

import pandas as pd

from datasets.models import FileType
from profiling.services import (
    PandasTabularReader,
    backfill_amount_column_from_qty_unit,
    detect_excel_header_row,
    is_structure_ambiguous,
    read_raw_tabular_grid,
)


def _build_excel(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales_List"
    ws["A1"] = "営業担当者100名分 販売実績一覧（仮想データ）"
    ws.merge_cells("A1:L1")
    ws["A2"] = "作成日"
    ws["B2"] = "2026-03-27"
    ws["D2"] = "対象期間"
    ws["E2"] = "2025-04-01 ～ 2026-03-31"
    ws["H2"] = "明細件数"
    ws["I2"] = "2,400"
    headers = [
        "担当者ID",
        "営業担当者名",
        "所属部門",
        "販売日",
        "顧客名",
        "都道府県",
        "商品カテゴリ",
        "商品名",
        "販売数量",
        "単価(円)",
        "売上金額(円)",
        "備考",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(4, i).value = h
    ws.cell(5, 1).value = "SR001"
    ws.cell(5, 2).value = "営業担当001"
    ws.cell(5, 4).value = "2025-04-18"
    ws.cell(5, 11).value = "416400"
    wb.save(path)


def test_detect_excel_header_row_for_title_and_meta_rows(tmp_path):
    p = tmp_path / "sample.xlsx"
    _build_excel(p)
    row = detect_excel_header_row(p, "Sales_List")
    assert row == 4


def test_reader_uses_detected_header_row(tmp_path):
    p = tmp_path / "sample.xlsx"
    _build_excel(p)
    reader = PandasTabularReader()
    df = reader.read_dataframe(p, FileType.XLSX, "Sales_List")
    assert "担当者ID" in list(df.columns)
    assert "Unnamed: 1" not in list(df.columns)


def test_read_raw_tabular_grid_matches_file_order(tmp_path):
    p = tmp_path / "sample.xlsx"
    _build_excel(p)
    grid, ncols = read_raw_tabular_grid(p, FileType.XLSX, "Sales_List", nrows=6)
    assert ncols >= 4
    assert "営業担当者100名分" in (grid[0][0] or "")
    assert grid[3][0] == "担当者ID"


def test_reader_respects_header_row_override(tmp_path):
    p = tmp_path / "sample.xlsx"
    _build_excel(p)
    reader = PandasTabularReader()
    df = reader.read_dataframe(p, FileType.XLSX, "Sales_List", header_row_1based=4)
    assert "担当者ID" in list(df.columns)


def test_is_structure_ambiguous_merged_cells():
    df = pd.DataFrame({"a": [1]})
    amb, reasons = is_structure_ambiguous({"merged_cells_count": 1}, df, FileType.XLSX)
    assert amb is True
    assert "merged_cells" in reasons


def test_is_structure_ambiguous_clean():
    df = pd.DataFrame({"担当者ID": ["x"], "売上": [1]})
    amb, _ = is_structure_ambiguous({"merged_cells_count": 0, "blank_row_ratio": 0.0}, df, FileType.XLSX)
    assert amb is False


def test_backfill_amount_when_formula_column_all_nan():
    df = pd.DataFrame(
        {
            "販売数量": ["2", "3"],
            "単価(円)": ["100", "200"],
            "売上金額(円)": [float("nan"), float("nan")],
        }
    )
    out = backfill_amount_column_from_qty_unit(df)
    assert out["売上金額(円)"].tolist() == ["200", "600"]
